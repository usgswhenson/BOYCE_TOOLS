'''
DEVELOPED BY SCOTT E BOYCE
CONTACT AT seboyce@usgs.gov or Boyce@engineer.com
CODE DEVELOMENT VERSION 1.0.0

Script loads HOB output and processes it for easy import in Excell
or makes a pdf document of plots from HOB.

Hob_To_Hydrograph_Excel_Chart.py  HOBFILE  EXCEL  SHEET_IN  SHEET_OUT  [HDRY]  [DYEAR] [CHART] [CHART_Cor]

    HOBFILE    => Hob File that will be processed.

    EXCEL      => Location of the Excel workbook that contains the input information.

    SHEET_IN   => Sheet that contains the input information.

    SHEET_OUT  => Sheet that will dump the output from the Hob File.
                  Set to SKIP to not add data to the sheet out.

    HDRY       => Any head value near this will be removed.
                  Set to NULL or SKIP to not use.

    DYEAR     => Optional, Default to False. Set to True if you want to use
                 decimal years in stead of dates in excel sheet

    CHART     => Optional, when present indicates the name the sheet that Excel Charts
                 should be placed in. If not present, then the chart sheet will not be made.

    CHART_Cor => Optional, when present indicates the name the sheet that Excel Correlation Charts
                 should be placed in. If not present, then the chart sheet will not be made.
                 You must specify CHART if you wish to make the CHART_Cor as well.

INPUTFILE Structure:
# Any line that starts with a # is skipped. Anything to the left of a # is ignored
OBSNAME   [GROUP]  [DD]  [SPREAD]  [DATE_START  DATE_END]

    OBSNAME    =>  Unique name contained within the Hob OBSname. This groups observation times.
                   Any Observation not specified here is skipped.

    GROUP      =>  Optional and when there is more than 1 group specified will lump observations into
                   one figure, so all observations appear side by side

    DD         =>  Overrules global drawdown/head specification for the specific observation. Must be "DD" or "HD".  (Global is not required to be specified.)

    SPREAD     =>  Overrules global SPREAD setting. (Global is not required to be specified.)
                   Set to "NONE" to use Global Setting (allows place holder to only specify dates)

    DATE_START =>  Overrules global DATE_START setting. (Global is not required to be specified.)

    DATE_END   =>  Overrules global DATE_END setting. (Global is not required to be specified.)


'''
#import openpyxl
import sys
import datetime as dt
from collections import OrderedDict
from copy import deepcopy as COPY
import numpy as np
#
from openpyxl       import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter as GL
from openpyxl.utils.units import points_to_pixels, pixels_to_EMU
from openpyxl.utils.datetime import to_excel as EXCEL_DATE
from openpyxl.styles import Font, Alignment
#from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice

#   SET CORRELATION MARKER AND SIZE. 
#
#        SIZE MUST BE >2.0  -- default size is 5
#
# CORR_MARKER = 'auto', 'plus', 'x', 'circle', 'diamond', 'square', 'star', 'triangle', 'dash', 'dot'
#
CORR_MARKER_ALL = 'circle'
CORR_SIZE_ALL   = 3

CORR_MARKER     = 'x'
CORR_SIZE       = 5

#CORRELATION COLOR, SET TO 'auto' to let Excel set automatically. It must be one of the colors in COLOR LIST:
CORR_COLOR_ALL = 'auto'
CORR_COLOR     = 'silver'

# COLOR LIST:
#
#  black                 gray                  olive                 silver                red                   blue
#  orange                brown                 green                 magenta               turquoise             gold
#  beige                 forestGreen           dkGoldenrod           lightPink             slateGrey             violet
#  white                 indigo                dkBlue                navajoWhite           lightCyan             chocolate
#  blanchedAlmond        lightCoral            seaGreen              yellow                limeGreen             ivory
#  dkSalmon              honeydew              lightSteelBlue        lavender              thistle               
#  ltCyan                lightGreen            lightSeaGreen         dkRed                 firebrick             
#  salmon                plum                  mediumTurquoise       lightGoldenrodYellow  peru                  
#  darkGrey              ltGray                lightYellow           dkCyan                orangeRed             
#  rosyBrown             paleGreen             darkMagenta           ltSlateGrey           burlyWood             
#  dodgerBlue            mediumSlateBlue       paleTurquoise         mediumVioletRed       floralWhite           
#  peachPuff             snow                  skyBlue               mediumPurple          ltSalmon              
#  ltGrey                medOrchid             medTurquoise          darkSlateBlue         darkTurquoise         
#  medSeaGreen           mediumOrchid          dkKhaki               tomato                darkBlue              
#  mediumBlue            darkKhaki             lemonChiffon          khaki                 ltSkyBlue             
#  cyan                  paleVioletRed         slateBlue             darkSalmon            dkOrchid              
#  medAquamarine         dimGrey               darkOrchid            lightSlateGray        medBlue               
#  dkSlateGrey           greenYellow           mintCream             dkTurquoise           deepPink              
#  dkSeaGreen            steelBlue             ltPink                darkSlateGrey         royalBlue             
#  teal                  medSlateBlue          lawnGreen             medVioletRed          ltSteelBlue           
#  seaShell              crimson               darkOliveGreen        ltBlue                medSpringGreen        
#  cadetBlue             papayaWhip            dimGray               pink                  mediumSeaGreen        
#  indianRed             lightBlue             dkOrange              maroon                goldenrod             
#  ltCoral               tan                   darkRed               deepSkyBlue           dkSlateGray           
#  azure                 whiteSmoke            chartreuse            darkCyan              mistyRose             
#  medPurple             sandyBrown            dkGrey                darkOrange            ltSeaGreen            
#  dkOliveGreen          moccasin              oldLace               lavenderBlush         darkGreen             
#  darkViolet            lightSalmon           cornsilk              linen                 hotPink               
#  yellowGreen           aliceBlue             gainsboro             ltYellow              orchid                
#  dkGreen               dkGray                navy                  darkGray              grey                  
#  ltSlateGray           ltGreen               paleGoldenrod         darkSlateGray         lightGray             
#  saddleBrown           lightGrey             slateGray             lightSkyBlue          springGreen           
#  bisque                midnightBlue          aqua                  powderBlue            mediumSpringGreen     
#  lime                  antiqueWhite          fuchsia               darkSeaGreen          purple                
#  cornflowerBlue        lightSlateGrey        dkSlateBlue           darkGoldenrod         oliveDrab             
#  dkViolet              mediumAquamarine      coral                 ghostWhite            aquamarine            
#  ltGoldenrodYellow     blueViolet            sienna                wheat                 dkMagenta             


#########################################################################################
#########################################################################################
#########################################################################################
def DYEAR_TO_DATE(DYEAR, USE_LEAP=True):
    # Given a decimal year (DYEAR) return datetime object with calendar date
    # Optional variable, USE_LEAP, is set to True  to use 365 and 366 day years
    #                                     or False to use  365.2425 days in a year
    YEAR = int(DYEAR)
    YDEC = DYEAR - float(YEAR)
    #
    if USE_LEAP:
        if (YEAR%4==0 and YEAR%100!=0) or YEAR%400:     # calendar.isleap(YEAR):
           CVRT = 366.
        else:
           CVRT = 365.
    else:
           CVRT = 365.2425
           #
    DoY = int( YDEC * CVRT  + 1.000001 )
    #
    FMT='%Y,%j'
    FDT=str(int(YEAR)) +','+str(int(DoY))
     #
    return dt.datetime.strptime(FDT,FMT).date()

def DATE_TO_DYEAR(DATE, USE_LEAP=True, FRAC=0.5):
   #Input can be either a list/Tuple that is structured as [Year, Month, Day]
   #Or as a datetime.date object.
   #
   if type(DATE)==tuple or type(DATE)==list:  DATE=dt.date(DATE[0],DATE[1],DATE[2])
   #
   FRAC = 1.0-FRAC
   YEAR = DATE.year
   #
   if not USE_LEAP:
        return YEAR + (float(DATE.strftime("%j"))-FRAC)/365.2425
   elif (YEAR%4==0 and YEAR%100!=0) or YEAR%400:   # calendar.isleap(DATE.year):
        return YEAR + (float(DATE.strftime("%j"))-FRAC)/366.0
   else:
        return YEAR + (float(DATE.strftime("%j"))-FRAC)/365.0

#########################################################################################
#########################################################################################
#########################################################################################

HOBFILE = sys.argv[1]
EXCEL   = sys.argv[2]
SHEET_IN= sys.argv[3]
SHEET_OUT= sys.argv[4]

try:
    HDRY = float(sys.argv[5])
except:
    HDRY = ''

try:
    DYEAR = sys.argv[6].upper()[0] in ['Y', 'T']
except:
    DYEAR = False

try:
    SHEET_OUT_CHART = sys.argv[7]
except:
    SHEET_OUT_CHART = SHEET_OUT+' Charts'

try:
    SHEET_OUT_CHART_CORR = sys.argv[8]
except:
    SHEET_OUT_CHART_CORR = SHEET_OUT_CHART+'_Corr'
try:
    LOGY = sys.argv[9]
except:
    LOGY = False
SHEET_OUT_CHART = unicode(SHEET_OUT_CHART)
SHEET_OUT_CHART_CORR = unicode(SHEET_OUT_CHART_CORR)

ID='BASE_ID'
GP='GROUP'
DDHD='DD/HD'
RST='RANGE_START'
REN='RANGE_END'
DST='DATE_START'
DEN='DATE_END'
CST='CORR_RANGE_START'
CEN='CORR_RANGE_END'
TT='TITLE'

wb = load_workbook(EXCEL, data_only=True, read_only=True)
ws = wb[SHEET_IN]

HEADER = [str(cell.value).upper() for cell in ws.rows.next()]

C=dict(zip([ID,GP,DDHD,RST,REN,DST,DEN,TT,CST,CEN],[-1]*10))
for I,HED in enumerate(HEADER):
    if   HED==ID:
                C[ID  ] = I
    elif HED==GP:
                C[GP  ] = I
    elif HED==DDHD:
                C[DDHD] = I
    elif HED==RST:
                C[RST ] = I
    elif HED==REN:
                C[REN ] = I
    elif HED==DST:
                C[DST ] = I
    elif HED==DEN:
                C[DEN ] = I
    elif HED==CST:
                C[CST ] = I
    elif HED==CEN:
                C[CEN ] = I
    elif HED==TT:
                C[TT  ] = I

GLOBAL_SPREAD = [None]
DATE_START = None
GLOBAL_DD  = False

NAMES=[]
GROUP=[]
DAT=OrderedDict()
GRP=OrderedDict()
SPREAD=[]
SPREAD_COR=[]
DD=[]
DSTR=[]
DEND=[]
TITLE=[]

for ROW in ws.iter_rows(row_offset=1):
    ROW =[ cell.value
           if not (unicode(cell.value).strip().upper() in ['NAN','NUL','NULL']) else None
           for cell in ROW ]
    #
    if (ROW[C[ID]] is None): continue
    #
    NAM = unicode(ROW[C[ID]])
    NAMES.append(NAM)
    #
    if not (NAM in DAT): DAT[NAM]=[]
    #
    if C[GP  ] >=0 and not (NAM is None) and not (ROW[C[GP]] is None):
        GROUP.append(unicode(ROW[C[GP]]))
        KEY=GROUP[-1]
        if not (KEY in GRP): GRP[KEY]=[]
        GRP[KEY].append(NAM)
    else:
        KEY=''
        if not (KEY in GRP): GRP[KEY]=[]
        GRP[KEY].append(NAM)
    #
    if C[DDHD ] >=0 and not (NAM is None) and not (ROW[C[DDHD]] is None):
        DD.append( str(ROW[C[DDHD]])[0].upper() in ['D','Y'] )
    else:
        DD.append(False)
    #
    if C[RST ] >=0 and not (NAM is None) and not (ROW[C[RST]] is None):
        SPREAD.append( [ ROW[C[RST]] ] )
        #
        if C[REN ] >=0 and not (ROW[C[REN]] is None):
            SPREAD[-1].append( ROW[C[REN]] )
    else:
        SPREAD.append( [None] )
    #
    if C[CST ] >=0 and not (NAM is None) and not (ROW[C[CST]] is None):
        SPREAD_COR.append( [ ROW[C[CST]] ] )
        #
        if C[CEN ] >=0  and not (ROW[C[CEN]] is None):
            SPREAD_COR[-1].append( ROW[C[CEN]] )
    else:
        SPREAD_COR.append( COPY(SPREAD[-1]) )
    #
    if C[DST ] >=0 and not (NAM is None) and not (ROW[C[DST]] is None):
        DT=ROW[C[DST]]
        if not DYEAR and type(DT) in [float,           int]: DT=DYEAR_TO_DATE(DT)
        if     DYEAR and type(DT) in [dt.datetime, dt.date]: DT=DATE_TO_DYEAR(DT)
        DSTR.append(DT)
    else:
        DSTR.append(None)
    #
    if C[DEN ] >=0 and not (NAM is None) and not (ROW[C[DEN]] is None):
        DT=ROW[C[DEN]]
        if not DYEAR and type(DT) in [float,           int]: DT=DYEAR_TO_DATE(DT)
        if     DYEAR and type(DT) in [dt.datetime, dt.date]: DT=DATE_TO_DYEAR(DT)
        DEND.append(DT)
    else:
        DEND.append(None)
    #
    if C[TT] >=0 and not (NAM is None) and not (ROW[C[TT]] is None):
        TITLE.append(unicode(ROW[C[TT]]))
    else:
        TITLE.append(NAM)
        #TITLE.append(None)

SPREAD= dict(zip(NAMES,SPREAD))
DD    = dict(zip(NAMES,DD))
DSTR  = dict(zip(NAMES,DSTR))
DEND  = dict(zip(NAMES,DEND))
TITLE = dict(zip(NAMES,TITLE))
SPREAD_COR= dict(zip(NAMES,SPREAD_COR))

wb._archive.close()
del ws, wb

####################################################################

CHECK = map(unicode.upper,NAMES)
ln=''
for I,NAM in enumerate(CHECK):
    N=len(NAM)
    for CHK in CHECK:
        if len(CHK)>N and NAM==CHK[:N]: ln+=NAM + ' and ' + CHK+'\n'
if ln !='':
    print ('ERROR: The following basenames has the first as a subset of the second.\nThis will cause confusion with the program identifying where the HOB observation is located at.\nThe following are the problem pairs:\n'+ln)
    sys.exit()
else:
    del CHECK

####################################################################

with open(HOBFILE,'r') as F:
    F.next()
    N = len(F.next().split())
    if N>3:
        INCLUDE_DATES=True
    else:
        INCLUDE_DATES=False
    F.seek(0)
    F.next()
    for ln in F:
        if ln.strip()=='':  continue
        ln=ln.split()
        SIM = ln[0]
        OBS = ln[1]
        NAM = ln[2].upper()
        #
        if HDRY!='' and abs( HDRY-float(SIM) )<1E-15: continue
        #
        if INCLUDE_DATES:
            DTs=ln[3]
            try:
                DT = dt.datetime.strptime(ln[3],'%Y-%m-%d').date()
            except:
                DT = dt.datetime.strptime(ln[3],'%m/%d/%Y').date()
            DYR= float(ln[4])
        for KEY in NAMES:
            N=len(KEY)
            if len(NAM)>=N and NAM[:N]==KEY:
                NAM='FOUND'
                break
        #
        if NAM != 'FOUND': continue
        #
        ln=[SIM,OBS]
        if INCLUDE_DATES:ln += [DTs,DT,DYR]
        DAT[KEY].append(ln)
ln=''
EMPTY=[]
for NAM in DAT:
    if len(DAT[NAM])==0:
        ln+=NAM+'\n'
        EMPTY.append(NAM)
        #del DAT[NAM]
        #NAMES.remove(NAM)
        #for KEY in GRP:
        #    if NAM in GRP[KEY]: GRP[KEY].remove(NAM)
if ln !='':
    print ('\n\nWARNING: The following basenames were not found in the HOB input file or contained all HDRY values.\n'+ln)

####################################################################

def Set_Header(SH,KEY,CC):
    SH.merge_cells(start_row=2,start_column=CC+1,end_row=2,end_column=CC+6)
    HED=SH.cell(row=2, column=CC+1)
    HED.value=KEY
    HED.alignment=Alignment(horizontal='center')
    HED.font = Font(name='Cambria', size=18, bold=True)

def SIZE(x):  #24pnt per row and 80 pnt per col
    return pixels_to_EMU(points_to_pixels(x))

#def PNT2CM(x):
#    return points_to_pixels(x)*0.02645833333  #cm/px @96dpi

def PX2CM(x):
    ppi = 96.0/2.54  #96 or 72
    return x/ppi  #cm/px @96dpi

def Set_BaseChart(CH):
    CH.style=None
    CH.x_axis.axPos='b'
    CH.x_axis.majorGridlines = None
   #CH.y_axis.majorGridlines = None
    #
    CH.legend.position = 'tr'  #['r', 'b', 'tr', 'l', 't']
    CH.legend.overlay=True
    #
    CH.height= 16*0.5085    # 1.018 => 2 Excel Rows
    CH.width =  8*1.694     #80pt/row v2.4 has 1.7111111111 = 1col # 7*2.202   # 2.202 => 1 column => 104 pt
#14.4/24

def AutoCorrSet(SH,R,C,MIN,MAX):
    RNG=MAX-MIN;    RNG=abs(RNG*0.1)
    MAX+=RNG;    MIN-=RNG
    SH.cell(row=R,   column=C, value=MIN);    SH.cell(row=R,   column=C+1, value=MIN)
    SH.cell(row=R+1, column=C, value=MAX);    SH.cell(row=R+1, column=C+1, value=MAX)
    X = Reference(SH, min_col=C,   min_row=R, max_row=R+1)
    Y = Reference(SH, min_col=C+1, min_row=R, max_row=R+1)
    return Series(X, Y, title='Autocorrelation Line')


#def MARKER(I):
#    #openpyxl.chart.marker
#    return ['picture', 'diamond', 'star', 'auto', 'dash', 'square', 'plus', 'x', 'triangle', 'circle', 'dot'][I]

from math import ceil, floor, log10
def CLEAN_RANGE(MIN,MAX):
    #
    RANGE = abs( MAX-MIN )
    if RANGE> 1E-20:
        DIG=int(log10(RANGE))
        FRC=log10(RANGE)-DIG
    else:
        return None, None
    #
    #if DIG<0.0: DIG+=1
    if   DIG>0 and FRC<0.69898:  #Number is less than half way to the next digit
                 DIG-=1
    elif DIG<0:
                 DIG+=1
    #
    RANGE=10**DIG
    #
    MIN = float( int(floor(MIN/RANGE)) * RANGE )
    MAX = float( int(ceil( MAX/RANGE)) * RANGE )
    #
    return MIN, MAX

if not (SHEET_OUT.upper() in ['SKIP','NONE']):
    #
    wb = load_workbook(EXCEL)
    #
    SHEET_OUT = unicode(SHEET_OUT)
    #
    if SHEET_OUT in wb.sheetnames:
        ws = wb[SHEET_OUT]
    else:
        ws = wb.create_sheet(title=SHEET_OUT)
    #
    if SHEET_OUT_CHART      in wb.sheetnames: wb.remove_sheet(wb[SHEET_OUT_CHART])
    if SHEET_OUT_CHART_CORR in wb.sheetnames: wb.remove_sheet(wb[SHEET_OUT_CHART_CORR])
    #
    IDX=wb.sheetnames.index(SHEET_OUT) + 1
    wb.create_sheet(index=IDX, title=SHEET_OUT_CHART)
    wb.create_sheet(index=IDX, title=SHEET_OUT_CHART_CORR)
    #
    wsch=wb[SHEET_OUT_CHART]
    wscr=wb[SHEET_OUT_CHART_CORR]
    #
    #wscr['C5']=-100000;    wscr['D5']=-100000
    #wscr['C6']= 100000;    wscr['D6']= 100000
    #X = Reference(wscr, min_col=3, min_row=5, max_row=6)
    #Y = Reference(wscr, min_col=4, min_row=5, max_row=6)
    #AutoCorr = Series(X, Y, title='Autocorrelation Line')
    #AutoCorr.graphicalProperties.line.solidFill = ColorChoice(prstClr='black')
    ##AutoCorr.graphicalProperties.solidFill=ColorChoice(prstClr='black')
    #AutoCorr.graphicalProperties.line.width = SIZE(1)         #SET TO 1 pnt = 19050
    #del X,Y
    #
    BASE_CHART = ScatterChart()
    Set_BaseChart(BASE_CHART)

#    BASE_CHART.y_axis.title = 'Head'
#    if DYEAR:
#        BASE_CHART.x_axis.title = 'Year'
#    else:
#        BASE_CHART.x_axis.title = 'Date'
    #
    BASE_CORR = ScatterChart()  #scatterStyle="marker"
    Set_BaseChart(BASE_CORR)
    BASE_CORR.legend.position = 'r'
    #BASE_CORR.width=BASE_CORR.height  #make it square
    #
    MAX_ROW=ws.max_row
    MAX_COL=ws.max_column
    if MAX_COL>len(NAMES)*4: MAX_COL = len(NAMES)*4
    #BLNK = ['' for I in range(MAX_COL)]
    #Blank out Worksheet
    JJ=range(1,MAX_COL+1)
    for I,R in enumerate(ws.iter_rows(),1):
        for J in JJ:
            ws.cell(row=I,column=J).value=None  #blank out sheet  _ is throw away variable
    del JJ
    #
    I=1
    ROW=[]
    for KEY in GRP:
            ROW+=[KEY,'','',''] + ['','','','']*(len(GRP[KEY])-1)
    for J,R in enumerate(ROW,1):
        ws.cell(row=I,column=J).value=R
    #
    I+=1
    ROW=[]
    for KEY in GRP:
        for NAM in GRP[KEY]:
            ROW+=[TITLE[NAM],'','','']
    for J,R in enumerate(ROW,1):
        ws.cell(row=I,column=J).value=R

    I+=1
    ROW=[]
    for KEY in GRP:
        for NAM in GRP[KEY]:
            ROW+=['Date',NAM+'_Sim',NAM+'_Obs',NAM+'_Res']
    for J,R in enumerate(ROW,1):
        ws.cell(row=I,column=J).value=R
    #
    ISTR=I+1
    J=-3
    CC=-8     #CHART COLUMN Start on Col 2
    for KEY in GRP:
        CR=ISTR  #CHART ROW
        CC+=10
        #
        Set_Header(wsch,KEY,CC)
        Set_Header(wscr,KEY,CC)
        CORR_ALL = COPY(BASE_CORR)
        CORR_ALL.title=KEY
        MIN_ALL= 1E100
        MAX_ALL=-1E100
        #
        #
        for NAM in GRP[KEY]:
            I=ISTR
            J+=4
            #
            if not (NAM in EMPTY):
                SIM_START = float(DAT[NAM][0][0])
                OBS_START = float(DAT[NAM][0][1])
            #MIN= 1E100
            #MAX=-1E100
            MIN= 1E100
            MAX=-1E100
            COR_MIN= 1E100
            COR_MAX=-1E100
            NOT_FIRST=False
            for D in DAT[NAM]:
                SIM=float(D[0])
                OBS=float(D[1])
                if   DD[NAM] and NOT_FIRST:
                           SIM = SIM + SIM_START
                           OBS = OBS + OBS_START
                NOT_FIRST=True
                if INCLUDE_DATES:
                    if DYEAR:
                        ws.cell(row=I,column=J).value=D[4]
                        ws.cell(row=I,column=J).number_format = '0.000000'
                    else:
                        ws.cell(row=I,column=J).value=D[3]
                        ws.cell(row=I,column=J).number_format = 'm/d/yyyy'
                else:
                    ws.cell(row=I,column=J).value=I-ISTR+1
                    ws.cell(row=I,column=J).number_format = 'General'
                J+=1
                ws.cell(row=I,column=J).value=SIM
                J+=1
                ws.cell(row=I,column=J).value=OBS
                J+=1
                ws.cell(row=I,column=J).value=OBS-SIM
                J-=3  #Move Back to start
                I+=1
                #
                if MIN>SIM: MIN=SIM
                if MIN>OBS: MIN=OBS
                if MAX<SIM: MAX=SIM
                if MAX<OBS: MAX=OBS
            #
            if MIN<MIN_ALL: MIN_ALL=MIN
            if MIN<MIN_ALL: MIN_ALL=MIN
            if MAX>MAX_ALL: MAX_ALL=MAX
            if MAX>MAX_ALL: MAX_ALL=MAX
            #
            if I<1000:
                I*=3
            else:
                I*=2
            CHART = COPY(BASE_CHART)
            Xval = Reference(ws, min_col=J,   min_row=ISTR, max_row=I)
            Ysim = Reference(ws, min_col=J+1, min_row=ISTR, max_row=I)
            Yobs = Reference(ws, min_col=J+2, min_row=ISTR, max_row=I)
            S = Series(Ysim, Xval, title='Sim')
            S.graphicalProperties.line.width = SIZE(1)         #SET TO 1 pnt = 19050   # 12700 EMU / 1 pt   19050 =>1.5 pt
            CHART.series.append(S)
            S = Series(Yobs, Xval, title='Obs')
            S.graphicalProperties.line.width = SIZE(1)         #SET TO 1 pnt = 19050   # 12700 EMU / 1 pt   19050 => 1.5pt
            CHART.series.append(S)
            CHART.title=TITLE[NAM]
            #
            if not (DSTR[NAM] is None) and not DYEAR: CHART.x_axis.scaling.min = EXCEL_DATE( DSTR[NAM] )
            if not (DEND[NAM] is None) and not DYEAR: CHART.x_axis.scaling.max = EXCEL_DATE( DEND[NAM] )
            #
            if not (DSTR[NAM] is None) and     DYEAR: CHART.x_axis.scaling.min = DSTR[NAM]
            if not (DEND[NAM] is None) and     DYEAR: CHART.x_axis.scaling.max = DEND[NAM]
            #
            if DYEAR:
                CHART.x_axis.number_format = '0.0'
            else:
                CHART.x_axis.number_format = 'm/YYYY'
            #
            if not (SPREAD[NAM][0] is None): #.graphicalProperties.line.
                if len(SPREAD[NAM]) == 1:
                    med = np.median(SIM)
                    CHART.y_axis.scaling.min = med - 0.5*SPREAD[NAM][0]
                    CHART.y_axis.scaling.max = med + 0.5*SPREAD[NAM][0]
                else:
                    CHART.y_axis.scaling.min = SPREAD[NAM][0]
                    CHART.y_axis.scaling.max = SPREAD[NAM][1]
            #
            CHART.x_axis.crossesAt=CHART.y_axis.scaling.min
            if LOGY:
                CHART.y_axis.scaling.logBase = 10
            #
            LOC=GL(CC)+str(CR)
            wsch.add_chart(CHART,LOC)
            #
            CORR = COPY(BASE_CORR)
            CORR.legend = None
            if (SPREAD_COR[NAM][0] is None): #.graphicalProperties.line.
                #x_axis_min = CHART.y_axis.scaling.min
                #x_axis_max = CHART.y_axis.scaling.max
                #y_axis_min = CHART.y_axis.scaling.min
                #y_axis_max = CHART.y_axis.scaling.max
                x_axis_min = MIN - (abs(MIN) * 0.01)
                x_axis_max = MAX + (abs(MAX) * 0.01)
                y_axis_min = x_axis_min
                y_axis_max = x_axis_max
            elif len(SPREAD_COR[NAM]) == 1:
                med = np.median(SIM)
                x_axis_min = med - 0.5*SPREAD_COR[NAM][0]
                x_axis_max = med + 0.5*SPREAD_COR[NAM][0]
                #
                y_axis_min = x_axis_min
                y_axis_max = x_axis_max
            else:
                x_axis_min = SPREAD_COR[NAM][0]
                x_axis_max = SPREAD_COR[NAM][1]
                y_axis_min = SPREAD_COR[NAM][0]
                y_axis_max = SPREAD_COR[NAM][1]
            #
            S = Series(Ysim, Yobs, title='Obs vs Sim')
            S.marker.size = float(CORR_SIZE)
            S.marker.symbol = CORR_MARKER
            #
            if CORR_COLOR != 'auto': 
                if CORR_MARKER in ['plus', 'x', 'star', 'dash']:
                    S.marker.graphicalProperties.line.solidFill = ColorChoice(prstClr=CORR_COLOR)
                    S.marker.graphicalProperties.noFill = True
                else:
                    S.marker.graphicalProperties.line.noFill = True
                    S.marker.graphicalProperties.solidFill = ColorChoice(prstClr=CORR_COLOR)
            #
            S.graphicalProperties.line.noFill = True
            CORR.series.append(S)
            #AutoCorr=AutoCorrSet(wscr,CR+19,CC+1,MIN,MAX)
            AutoCorr=AutoCorrSet(wscr,CR+19,CC+1, x_axis_min-abs(x_axis_min),y_axis_min+abs(y_axis_min))
            AutoCorr.graphicalProperties.line.solidFill = ColorChoice(prstClr='black')
            AutoCorr.graphicalProperties.solidFill=ColorChoice(prstClr='black')
            AutoCorr.graphicalProperties.line.width = SIZE(1)         #SET TO 1 pnt = 19050
            CORR.series.append( AutoCorr )
            #
            CORR.title=TITLE[NAM]
            #CORR.x_axis.title='Observed'
            #CORR.y_axis.title='Simulated'
            #CORR.x_axis.title.text.rich.paragraphs[0].r.rPr=openpyxl.drawing.text.CharacterProperties()
            #CORR.x_axis.title.text.rich.paragraphs[0].r.properties.sz=12
            #CORR.x_axis.title.text.rich.paragraphs[0].r.rPr.cs=Font(name='Cambria', size=12, bold=True)
            #
            ##CORR.x_axis.scaling.min = CHART.y_axis.scaling.min
            ##CORR.x_axis.scaling.max = CHART.y_axis.scaling.max
            ##CORR.y_axis.scaling.min = CHART.y_axis.scaling.min
            ##CORR.y_axis.scaling.max = CHART.y_axis.scaling.max
            #
            #CORR.x_axis.scaling.min = MIN_O
            #CORR.x_axis.scaling.max = MAX_O
            #CORR.y_axis.scaling.min = MIN_S
            #CORR.y_axis.scaling.max = MAX_S
            #
            #CORR.x_axis.crossesAt=CORR.y_axis.scaling.min
            #CORR.y_axis.crossesAt=CORR.x_axis.scaling.min
            #
            CORR.x_axis.scaling.min = x_axis_min
            CORR.x_axis.scaling.max = x_axis_max
            CORR.y_axis.scaling.min = y_axis_min
            CORR.y_axis.scaling.max = y_axis_max
            #
            COR_MIN=min(COR_MIN, x_axis_min)
            COR_MAX=max(COR_MAX, x_axis_max)
            #
            CORR.x_axis.crosses='min'
            CORR.y_axis.crosses='min'
            #
            CORR.x_axis.number_format = 'General'
            CORR.y_axis.number_format = 'General'
            #
            LOC=GL(CC)+str(CR+18)
            wscr.add_chart(CORR,LOC)
            #
            S = Series(Ysim, Yobs, title=TITLE[NAM])
            S.marker.size   = float(CORR_SIZE_ALL)
            S.marker.symbol = CORR_MARKER_ALL
            S.graphicalProperties.line.noFill = True
            #
            if CORR_COLOR_ALL != 'auto': 
                if CORR_MARKER_ALL in ['plus', 'x', 'star', 'dash']:
                    S.marker.graphicalProperties.line.solidFill = ColorChoice(prstClr=CORR_COLOR_ALL)
                    S.marker.graphicalProperties.noFill = True
                else:
                    S.marker.graphicalProperties.line.noFill = True
                    S.marker.graphicalProperties.solidFill = ColorChoice(prstClr=CORR_COLOR_ALL)
            #
            CORR_ALL.append(S)
            #
            CR+=18
        #
        AutoCorr=AutoCorrSet(wscr,ISTR+1,CC+1,COR_MIN,COR_MAX)   #,MIN_ALL,MAX_ALL
        AutoCorr.graphicalProperties.line.solidFill = ColorChoice(prstClr='black')
        AutoCorr.graphicalProperties.solidFill=ColorChoice(prstClr='black')
        AutoCorr.graphicalProperties.line.width = SIZE(1)         #SET TO 1 pnt = 19050
        CORR_ALL.series.append(AutoCorr)
        CORR_ALL.x_axis.scaling.min = COR_MIN
        CORR_ALL.x_axis.scaling.max = COR_MAX
        CORR_ALL.y_axis.scaling.min = COR_MIN
        CORR_ALL.y_axis.scaling.max = COR_MAX
        #
        #CORR_ALL.x_axis.crossesAt=CORR_ALL.y_axis.scaling.min
        #CORR_ALL.y_axis.crossesAt=CORR_ALL.x_axis.scaling.min
        #
        #CORR_ALL.x_axis.crossesAt=CHART.y_axis.scaling.min
        #
        CORR_ALL.x_axis.crosses='min'
        CORR_ALL.y_axis.crosses='min'
        #
        CORR_ALL.x_axis.number_format = 'General'
        CORR_ALL.y_axis.number_format = 'General'
        #
        CORR_ALL.x_axis.title='Observed'
        CORR_ALL.y_axis.title='Simulated'
        #
        LOC=GL(CC)+str(ISTR)
        wscr.add_chart(CORR_ALL,LOC)
        #
    #
    # TODO fix this
    # raw_input('\n\n        ...WARNING...\n\nExcel File about to be saved.\nMake sure that the workbook is closed.\nIf script is interupted it can corrupt the Excel File.\n\nPress Enter to Continue\n\n')
    wb.save(EXCEL)
    print ('...Excel Workbook Saved...\n\n')
